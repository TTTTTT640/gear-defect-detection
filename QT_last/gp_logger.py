"""
检测日志系统 - 记录每次检测结果，支持Excel导出
"""

import time
import threading
from datetime import datetime


class DetectionLogger:
    """检测日志管理器"""

    def __init__(self):
        self.records = []
        self.lock = threading.Lock()
        self._consecutive_defects = 0  # 连续缺陷计数(供SMS报警用)

    def log_detection(self, stage, class_name, confidence, serial_cmd, inference_ms=0):
        """记录一条检测结果

        Args:
            stage: 检测阶段 (0=存在性检测, 1=缺陷分类)
            class_name: 检测类别 (good/miss/bad/well)
            confidence: 置信度 (0~1)
            serial_cmd: 发送的串口指令 (01/02/07/08)
            inference_ms: 推理耗时(毫秒)
        """
        record = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "stage": stage,
            "class_name": class_name,
            "confidence": round(confidence, 4),
            "serial_cmd": serial_cmd,
            "action": self._cmd_to_action(serial_cmd),
            "inference_ms": round(inference_ms, 1),
        }
        with self.lock:
            self.records.append(record)
            # 更新连续缺陷计数
            if class_name in ("bad", "miss"):
                self._consecutive_defects += 1
            elif class_name in ("good", "well"):
                self._consecutive_defects = 0

    @property
    def consecutive_defects(self):
        with self.lock:
            return self._consecutive_defects

    def get_stats(self):
        """获取检测统计信息"""
        with self.lock:
            if not self.records:
                return {"total": 0, "defect_rate": 0.0}

            counts = {}
            for r in self.records:
                cn = r["class_name"]
                counts[cn] = counts.get(cn, 0) + 1

            total = len(self.records)
            defects = counts.get("bad", 0) + counts.get("miss", 0)
            return {
                "total": total,
                "good": counts.get("good", 0),
                "well": counts.get("well", 0),
                "bad": counts.get("bad", 0),
                "miss": counts.get("miss", 0),
                "defect_rate": round(defects / total * 100, 1) if total > 0 else 0.0,
                "pass_rate": round((total - defects) / total * 100, 1) if total > 0 else 0.0,
            }

    def export_excel(self, filepath):
        """导出检测日志到Excel文件

        Args:
            filepath: 输出文件路径 (.xlsx)
        Returns:
            bool: 是否成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("请安装 openpyxl: pip install openpyxl")
            return False

        wb = Workbook()

        # ---- Sheet 1: 检测明细 ----
        ws = wb.active
        ws.title = "检测明细"

        headers = ["时间", "阶段", "检测类别", "置信度", "串口指令", "动作", "推理耗时(ms)"]
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 条件格式颜色
        red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

        with self.lock:
            for row_idx, r in enumerate(self.records, 2):
                values = [
                    r["timestamp"],
                    f"阶段{r['stage']}",
                    r["class_name"],
                    f"{r['confidence']:.2%}",
                    r["serial_cmd"],
                    r["action"],
                    r["inference_ms"],
                ]
                fill = None
                if r["class_name"] in ("bad", "miss"):
                    fill = red_fill
                elif r["class_name"] in ("well", "good"):
                    fill = green_fill

                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=v)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if fill:
                        cell.fill = fill

        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # ---- Sheet 2: 统计摘要 ----
        ws2 = wb.create_sheet("统计摘要")
        stats = self.get_stats()
        summary = [
            ("检测总数", stats["total"]),
            ("合格(good)", stats.get("good", 0)),
            ("合格(well)", stats.get("well", 0)),
            ("缺陷(bad)", stats.get("bad", 0)),
            ("缺齿(miss)", stats.get("miss", 0)),
            ("缺陷率", f"{stats['defect_rate']}%"),
            ("合格率", f"{stats['pass_rate']}%"),
            ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for row_idx, (label, value) in enumerate(summary, 1):
            ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=row_idx, column=2, value=value)
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 20

        wb.save(filepath)
        print(f"日志已导出: {filepath}")
        return True

    def clear(self):
        """清空日志"""
        with self.lock:
            self.records.clear()
            self._consecutive_defects = 0

    @staticmethod
    def _cmd_to_action(cmd):
        return {
            "01": "合格放行",
            "02": "缺陷剔除",
            "07": "启动传送带",
            "08": "停止传送带",
        }.get(cmd, f"未知({cmd})")
